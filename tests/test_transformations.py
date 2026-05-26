from __future__ import annotations

import json
import re
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
import pandas as pd

from src.charts import monthly_bar
from src.insights import generate_executive_insights
from src.manager_exports import (
    build_executive_pdf,
    build_full_pdf,
    build_interactive_html_report,
    build_manager_excel,
    build_visuals_pdf,
    manager_chart_datasets,
    manager_dynamic_insights,
    manager_records_to_csv,
    manager_supplier_detail,
)
from src.metrics import compute_kpis
from src.metrics import compute_executive_metrics
from src.transformations import (
    INVOICE_COLUMNS,
    build_monthly_forecast,
    canonicalize_invoice_columns,
    clean_invoices,
    clean_monthly_summary,
    date_quality_summary,
    filter_invoices,
    parse_amount,
    parse_excel_date,
)
from src.ui import (
    _dataframe_to_excel_bytes,
    prepare_display_dataframe,
    prepare_forecast_export,
    prepare_invoice_display,
    prepare_invoice_export,
)


def make_invoice_frame() -> pd.DataFrame:
    rows = [
        ["Alpha SARL", "A-001", "15/01/2025", "1 200,50 MAD", "01/01/2025", "CA-1", "Projet A", "01"],
        ["Beta SA", "B-001", "20/01/2025", "300 DH", "02/01/2025", "CA-2", "Projet B", "01"],
        ["Alpha SARL", "A-002", "05/02/2025", "2 000", "03/01/2025", "CA-1", "Projet A", "02"],
    ]
    return pd.DataFrame(rows, columns=INVOICE_COLUMNS)


def extract_html_report_js(html: str) -> str:
    match = re.search(
        r'<script id="report-data" type="application/json">.*?</script>\s*<script>\s*(.*?)\s*</script>',
        html,
        flags=re.DOTALL,
    )
    assert match is not None
    return match.group(1)


def test_manager_exports_source_is_not_duplicated_or_mojibaked() -> None:
    source = Path("src/manager_exports.py").read_text(encoding="utf-8")
    function_names = [
        "_date_text",
        "_manager_records",
        "manager_chart_datasets",
        "manager_supplier_detail",
        "manager_dynamic_insights",
        "manager_records_to_csv",
        "_json_for_html_script",
        "format_filters_summary",
        "build_manager_excel",
        "_styles",
        "_chart_image",
        "_add_chart",
        "_table_from_dataframe",
        "_build_pdf",
        "build_executive_pdf",
        "build_visuals_pdf",
        "build_full_pdf",
    ]

    assert "Ã" not in source
    assert "Â" not in source
    assert "�" not in source
    assert source.count("from __future__ import annotations") == 1
    assert source.count("def build_interactive_html_report(") == 1
    assert source.count('html = """<!doctype html>') == 1
    assert ("Rapport " + "manager") not in source
    for name in function_names:
        assert source.count(f"def {name}(") == 1


def test_parse_amount_accepts_french_formats_and_currency_labels() -> None:
    values = pd.Series(["1 234,56 MAD", "2\u00a0500,00 DH", "300,5", "4 000"])

    parsed = parse_amount(values)

    assert parsed.tolist() == [1234.56, 2500.0, 300.5, 4000.0]


def test_parse_excel_serial_date() -> None:
    parsed = parse_excel_date(pd.Series([46139]))

    assert parsed.iloc[0] == pd.Timestamp("2026-04-27")


def test_parse_date_comptable_dot_format() -> None:
    parsed = parse_excel_date(pd.Series(["27.01.2026", "18.02.2026"]))

    assert parsed.iloc[0] == pd.Timestamp("2026-01-27")
    assert parsed.iloc[1] == pd.Timestamp("2026-02-18")


def test_parse_text_date_day_first() -> None:
    parsed = parse_excel_date(pd.Series(["15/02/2025"]))

    assert parsed.iloc[0] == pd.Timestamp("2025-02-15")


def test_parse_mixed_real_excel_date_values_without_nat_strings() -> None:
    values = pd.Series(
        [
            datetime(2025, 1, 15, 13, 30),
            pd.Timestamp("2025-02-20"),
            date(2025, 3, 5),
            45748,
            "30/04/2025",
            "2025-05-12",
            "",
            None,
        ]
    )

    parsed = parse_excel_date(values)

    assert parsed.iloc[0] == pd.Timestamp("2025-01-15")
    assert parsed.iloc[1] == pd.Timestamp("2025-02-20")
    assert parsed.iloc[2] == pd.Timestamp("2025-03-05")
    assert parsed.iloc[3] == pd.Timestamp("2025-04-01")
    assert parsed.iloc[4] == pd.Timestamp("2025-04-30")
    assert parsed.iloc[5] == pd.Timestamp("2025-05-12")
    assert parsed.iloc[6:].isna().all()


def test_missing_expected_invoice_columns_do_not_crash() -> None:
    raw = pd.DataFrame(
        {
            "Nom du fournisseur": ["Alpha SARL"],
            "Montant": ["1 000,00 MAD"],
            "Date d'échéance": ["15/01/2025"],
        }
    )

    cleaned = clean_invoices(raw)

    assert len(cleaned) == 1
    assert cleaned.loc[0, "Montant"] == 1000
    assert cleaned.loc[0, "Code Affaire"] == "Non renseigné"


def test_invalid_due_dates_do_not_leak_nat_or_none_to_month_year() -> None:
    raw = pd.DataFrame(
        {
            "Nom du fournisseur": ["Alpha SARL", "Beta SA"],
            "Montant": [1000, 2000],
            "Date d'échéance": ["", "not a date"],
        }
    )

    cleaned = clean_invoices(raw)
    quality = date_quality_summary(cleaned)
    forecast = build_monthly_forecast(cleaned)

    assert cleaned["Mois"].tolist() == ["Date invalide", "Date invalide"]
    assert cleaned["Année"].tolist() == ["Date invalide", "Date invalide"]
    assert "NaT" not in cleaned["Mois"].astype(str).tolist()
    assert quality["invalid_rows"] == 2
    assert forecast.empty


def test_prepare_display_dataframe_hides_nat_and_none_values() -> None:
    raw = pd.DataFrame(
        {
            "date": [pd.Timestamp("2025-01-01"), pd.NaT],
            "texte": ["ok", None],
        }
    )

    display = prepare_display_dataframe(raw)

    assert display.loc[0, "date"] == "01/01/2025"
    assert display.loc[1, "date"] == ""
    assert display.loc[1, "texte"] == ""


def test_invoice_display_has_no_internal_columns_or_nat_none() -> None:
    cleaned = clean_invoices(
        pd.DataFrame(
            {
                "Nom du fournisseur": ["Alpha SARL"],
                "Numéro réf. fournisseur": ["A-1"],
                "Date d'échéance": [pd.NaT],
                "Montant": [1000],
                "Date comptable": [None],
                "Code Affaire": ["CA-1"],
                "Nom Projet": ["Projet A"],
                "MOIS": ["04"],
            }
        )
    )

    display = prepare_invoice_display(cleaned)

    assert display.columns.tolist() == [
        "Nom du fournisseur",
        "Numéro réf. fournisseur",
        "Date d'échéance",
        "Montant",
        "Date comptable",
        "Code Affaire",
        "Nom Projet",
        "Mois source",
        "Mois calculé",
        "Année",
    ]
    assert "Date d'échéance valide" not in display.columns
    assert "NaT" not in display.to_string()
    assert "None" not in display.to_string()


def test_invoice_export_keeps_amount_numeric_and_dates_as_dates() -> None:
    cleaned = clean_invoices(make_invoice_frame())

    exported = prepare_invoice_export(cleaned)

    assert pd.api.types.is_numeric_dtype(exported["Montant"])
    assert pd.api.types.is_datetime64_any_dtype(exported["Date d'échéance"])
    assert pd.api.types.is_datetime64_any_dtype(exported["Date comptable"])
    assert pd.api.types.is_datetime64_any_dtype(exported["Mois calculé"])
    assert not exported["Montant"].astype(str).str.contains("MAD").any()


def test_invoice_excel_export_round_trips_dates_without_raw_serial_values() -> None:
    cleaned = clean_invoices(
        pd.DataFrame(
            {
                "Nom du fournisseur": ["Alpha SARL"],
                "Numéro réf. fournisseur": ["A-1"],
                "Date d'échéance": [46139],
                "Montant": [1000],
                "Date comptable": ["27.01.2026"],
                "Code Affaire": ["CA-1"],
                "Nom Projet": ["Projet A"],
                "MOIS": ["04"],
            }
        )
    )
    exported = prepare_invoice_export(cleaned)

    excel_bytes = _dataframe_to_excel_bytes({"Factures filtrées": exported})
    round_trip = pd.read_excel(BytesIO(excel_bytes), sheet_name="Factures filtrées")

    assert pd.api.types.is_datetime64_any_dtype(round_trip["Date d'échéance"])
    assert pd.api.types.is_datetime64_any_dtype(round_trip["Date comptable"])
    assert pd.api.types.is_datetime64_any_dtype(round_trip["Mois calculé"])
    assert round_trip.loc[0, "Date d'échéance"] == pd.Timestamp("2026-04-27")
    assert round_trip.loc[0, "Date comptable"] == pd.Timestamp("2026-01-27")
    assert round_trip.loc[0, "Mois calculé"] == pd.Timestamp("2026-04-01")
    assert round_trip.loc[0, "Date d'échéance"] != 46139


def test_invoice_export_converts_realistic_serial_values_before_excel_write() -> None:
    raw_export = pd.DataFrame(
        {
            "Nom du fournisseur": ["Alpha SARL"],
            "Numéro réf. fournisseur": ["A-1"],
            "Date d'échéance": [46139],
            "Montant": [1000.0],
            "Date comptable": [46049],
            "Code Affaire": ["CA-1"],
            "Nom Projet": ["Projet A"],
            "Mois source": ["04"],
            "Mois": [46113],
            "Année": ["2026"],
        }
    )

    exported = prepare_invoice_export(raw_export)

    assert exported.loc[0, "Date d'échéance"] == pd.Timestamp("2026-04-27")
    assert exported.loc[0, "Date comptable"] == pd.Timestamp("2026-01-27")
    assert exported.loc[0, "Mois calculé"] == pd.Timestamp("2026-04-01")
    assert pd.api.types.is_datetime64_any_dtype(exported["Date d'échéance"])
    assert pd.api.types.is_datetime64_any_dtype(exported["Date comptable"])
    assert pd.api.types.is_datetime64_any_dtype(exported["Mois calculé"])


def test_excel_export_writes_actual_date_cells_not_numeric_serial_cells() -> None:
    raw_export = pd.DataFrame(
        {
            "Nom du fournisseur": ["Alpha SARL"],
            "Numéro réf. fournisseur": ["A-1"],
            "Date d'échéance": [46139],
            "Montant": [1000.0],
            "Date comptable": [46049],
            "Code Affaire": ["CA-1"],
            "Nom Projet": ["Projet A"],
            "Mois source": ["04"],
            "Mois": [46113],
            "Année": ["2026"],
        }
    )
    exported = prepare_invoice_export(raw_export)
    excel_bytes = _dataframe_to_excel_bytes({"Factures filtrées": exported})

    workbook = load_workbook(BytesIO(excel_bytes), data_only=True)
    sheet = workbook["Factures filtrées"]
    headers = [cell.value for cell in sheet[1]]
    due_date_cell = sheet.cell(row=2, column=headers.index("Date d'échéance") + 1)
    accounting_date_cell = sheet.cell(row=2, column=headers.index("Date comptable") + 1)
    month_cell = sheet.cell(row=2, column=headers.index("Mois calculé") + 1)

    assert due_date_cell.is_date
    assert accounting_date_cell.is_date
    assert month_cell.is_date
    assert due_date_cell.value == datetime(2026, 4, 27)
    assert accounting_date_cell.value == datetime(2026, 1, 27)
    assert month_cell.value == datetime(2026, 4, 1)


def test_forecast_export_keeps_amounts_numeric_without_currency_text() -> None:
    cleaned = clean_invoices(make_invoice_frame())
    forecast = build_monthly_forecast(cleaned)

    exported = prepare_forecast_export(forecast)

    assert pd.api.types.is_numeric_dtype(exported["Montant"])
    assert pd.api.types.is_numeric_dtype(exported["Cumul"])
    assert not exported["Montant"].astype(str).str.contains("MAD").any()
    assert not exported["Cumul"].astype(str).str.contains("MAD").any()


def test_display_tables_can_still_show_currency_text() -> None:
    cleaned = clean_invoices(make_invoice_frame())

    display = prepare_invoice_display(cleaned)

    assert display["Montant"].astype(str).str.contains("MAD").all()


def test_column_matching_is_tolerant_and_avoids_duplicate_crashes() -> None:
    raw = pd.DataFrame(
        {
            "  NOM\nDU FOURNISSEUR  ": ["Alpha SARL"],
            "Numéro réf. Fournisseur": ["A-001"],
            "Numero ref fournisseur": [pd.NA],
            "Date echeance": [45658],
            "MONTANT": ["1 500,00 DH"],
            "MOIS": ["01"],
        }
    )

    cleaned = clean_invoices(raw)

    assert cleaned.loc[0, "Nom du fournisseur"] == "Alpha SARL"
    assert cleaned.loc[0, "Numéro réf. fournisseur"] == "A-001"
    assert cleaned.loc[0, "Mois"] == "2025-01"
    assert cleaned.loc[0, "Montant"] == 1500


def test_column_canonicalization_directly_maps_variants() -> None:
    raw = pd.DataFrame(
        {
            "Nom\n du Fournisseur": ["Alpha SARL"],
            "Numéro réf. Fournisseur": ["A-001"],
            "DATE ECHEANCE": ["15/01/2025"],
            " montant ": [100],
        }
    )

    canonical = canonicalize_invoice_columns(raw)

    assert "Nom du fournisseur" in canonical.columns
    assert "Numéro réf. fournisseur" in canonical.columns
    assert "Date d'échéance" in canonical.columns
    assert canonical.loc[0, "Montant"] == 100


def test_final_feuil1_schema_without_projet_partenaire_and_month_source() -> None:
    raw = pd.DataFrame(
        {
            "Nom du fournisseur": ["Alpha SARL"],
            "Numéro réf. fournisseur": [12345],
            "Date d'échéance": [46139],
            "Montant": [1000],
            "Date comptable": ["27.01.2026"],
            "Code Affaire": ["CA-9"],
            "Nom Projet": ["Projet Final"],
            "MOIS": [4],
        }
    )

    cleaned = clean_invoices(raw)

    assert cleaned.loc[0, "Date d'échéance"] == pd.Timestamp("2026-04-27")
    assert cleaned.loc[0, "Date comptable"] == pd.Timestamp("2026-01-27")
    assert cleaned.loc[0, "Mois source"] == "04"
    assert cleaned.loc[0, "Mois"] == "2026-04"
    assert cleaned.loc[0, "Projet / Code Affaire"] == "Projet Final / CA-9"


def test_feuil2_monthly_summary_keeps_retard_and_penalty_columns() -> None:
    summary = pd.DataFrame(
        {
            "mois": ["04"],
            "dépenses prévisionnelles": ["1 000,50"],
            "Dépenses réelles": [900],
            "ecarts": [100.5],
            "Retard en jours": [5],
            "Montant de pénalité": ["250 DH"],
        }
    )

    cleaned = clean_monthly_summary(summary)

    assert cleaned.loc[0, "Retard en jours"] == 5
    assert cleaned.loc[0, "Montant de pénalité"] == 250


def test_monthly_forecast_aggregation() -> None:
    cleaned = clean_invoices(make_invoice_frame())

    forecast = build_monthly_forecast(cleaned)

    january = forecast[forecast["Mois"] == "2025-01"].iloc[0]
    february = forecast[forecast["Mois"] == "2025-02"].iloc[0]
    assert january["Montant"] == 1500.5
    assert january["Nombre de factures"] == 2
    assert february["Montant"] == 2000
    assert february["Cumul"] == 3500.5


def test_compute_kpis_empty_and_non_empty() -> None:
    empty_kpis = compute_kpis(pd.DataFrame(), pd.DataFrame())
    assert empty_kpis["Total à payer"] == "0.00 MAD"
    assert empty_kpis["Nombre de factures"] == "0"

    cleaned = clean_invoices(make_invoice_frame())
    forecast = build_monthly_forecast(cleaned)
    kpis = compute_kpis(cleaned, forecast)

    assert kpis["Total à payer"] == "3 500.50 MAD"
    assert kpis["Nombre de factures"] == "3"
    assert kpis["Nombre de fournisseurs"] == "2"
    assert set(kpis) == {"Total à payer", "Nombre de factures", "Nombre de fournisseurs", "Mois le plus élevé"}


def test_risk_metric_calculations() -> None:
    cleaned = clean_invoices(make_invoice_frame())
    forecast = build_monthly_forecast(cleaned)

    metrics = compute_executive_metrics(cleaned, forecast)

    assert round(metrics["top5_supplier_share"], 4) == 1
    assert metrics["top5_supplier_severity"] == "Élevé"
    assert metrics["highest_month"]["label"] == "2025-02"
    assert metrics["average_invoice_amount"] == cleaned["Montant"].mean()


def test_executive_insight_generation() -> None:
    cleaned = clean_invoices(make_invoice_frame())
    forecast = build_monthly_forecast(cleaned)
    metrics = compute_executive_metrics(cleaned, forecast)

    insights = generate_executive_insights(cleaned, forecast, metrics)

    assert any("2025-02" in insight for insight in insights)
    assert any("5 principaux fournisseurs" in insight for insight in insights)
    assert any("Projet A / CA-1" in insight for insight in insights)


def test_empty_chart_handling() -> None:
    fig = monthly_bar(pd.DataFrame())

    assert len(fig.layout.annotations) == 1
    assert "Aucune échéance" in fig.layout.annotations[0].text


def test_filter_invoices_behavior() -> None:
    cleaned = clean_invoices(make_invoice_frame())

    filtered = filter_invoices(
        cleaned,
        suppliers=["Alpha SARL"],
        codes=["CA-1"],
        project_names=["Projet A"],
        months=["2025-02"],
        min_amount=1000,
    )

    assert len(filtered) == 1
    assert filtered.iloc[0]["Numéro réf. fournisseur"] == "A-002"


def test_manager_excel_export_respects_filtered_data() -> None:
    cleaned = clean_invoices(make_invoice_frame())
    filtered = filter_invoices(
        cleaned,
        suppliers=["Alpha SARL"],
        codes=[],
        project_names=[],
        months=["2025-02"],
        min_amount=0,
    )
    forecast = build_monthly_forecast(filtered)
    top = pd.DataFrame(
        {
            "Nom du fournisseur": ["Alpha SARL"],
            "Montant": [2000.0],
            "Nombre de factures": [1],
        }
    )

    content = build_manager_excel(filtered, forecast, top)
    workbook = pd.ExcelFile(BytesIO(content))
    invoices = pd.read_excel(workbook, sheet_name="Factures filtrées")
    exported_forecast = pd.read_excel(workbook, sheet_name="Prévision mensuelle")

    assert len(invoices) == 1
    assert invoices.loc[0, "Nom du fournisseur"] == "Alpha SARL"
    assert exported_forecast.loc[0, "Mois"] == "2025-02"
    assert exported_forecast.loc[0, "Montant"] == 2000


def test_manager_pdf_exports_are_non_empty() -> None:
    cleaned = clean_invoices(make_invoice_frame())
    forecast = build_monthly_forecast(cleaned)
    top = pd.DataFrame(
        {
            "Nom du fournisseur": ["Alpha SARL", "Beta SA"],
            "Montant": [3200.5, 300.0],
            "Nombre de factures": [2, 1],
        }
    )
    projects = pd.DataFrame({"Projet / Code Affaire": ["Projet A / CA-1"], "Montant": [3200.5]})
    pareto = pd.DataFrame(
        {
            "Nom du fournisseur": ["Alpha SARL", "Beta SA"],
            "Montant": [3200.5, 300.0],
            "Contribution": [0.91, 0.09],
            "Contribution cumulée": [0.91, 1.0],
        }
    )
    metrics = compute_executive_metrics(cleaned, forecast)
    kpis = compute_kpis(cleaned, forecast)
    insights = generate_executive_insights(cleaned, forecast, metrics)
    filters = {"suppliers": [], "codes": [], "project_names": [], "months": [], "min_amount": 0.0}

    summary = build_executive_pdf(filters, kpis, metrics, insights, forecast, top)
    visuals = build_visuals_pdf(forecast, top, projects, pareto)
    full = build_full_pdf(filters, kpis, metrics, insights, forecast, top, projects, pareto, cleaned)

    assert summary.startswith(b"%PDF") and len(summary) > 1000
    assert visuals.startswith(b"%PDF") and len(visuals) > 1000
    assert full.startswith(b"%PDF") and len(full) > 1000


def test_interactive_html_manager_report_contains_data_and_controls() -> None:
    cleaned = clean_invoices(make_invoice_frame())
    filtered = filter_invoices(
        cleaned,
        suppliers=["Alpha SARL"],
        codes=[],
        project_names=[],
        months=["2025-02"],
        min_amount=0,
    )
    forecast = build_monthly_forecast(filtered)
    metrics = compute_executive_metrics(filtered, forecast)
    kpis = compute_kpis(filtered, forecast)
    insights = generate_executive_insights(filtered, forecast, metrics)
    filters = {
        "suppliers": ["Alpha SARL"],
        "codes": [],
        "project_names": [],
        "months": ["2025-02"],
        "min_amount": 0.0,
    }

    content = build_interactive_html_report(filters, filtered, kpis, metrics, insights)
    html = content.decode("utf-8")
    js = extract_html_report_js(html)

    assert len(content) > 1000
    assert '<script id="report-data" type="application/json">' in html
    assert html.count('<script id="report-data" type="application/json">') == 1
    assert html.count("</script>\n<script>") == 1
    assert html.count("</main>") == 1
    assert html.count("</body></html>") == 1
    assert html.count("<title>") == 1
    assert html.count("<header>") == 1
    assert html.count("const el =") == 1
    assert html.count(":root{") == 1
    assert html.count(".brand{") == 1
    assert html.count("<h2>Factures filtrées</h2>") == 1
    assert html.count('id="tableSearch"') == 1
    assert html.count('id="tableChips"') == 1
    assert html.count('id="noResults"') == 1
    assert html.count('id="noResultsReset"') == 1
    assert html.count('id="detailTable"') == 1
    assert html.count('id="resetBtn"') == 1
    assert html.count('id="filterNotice"') == 1
    assert "Alpha SARL" in html
    assert "2025-02" in html
    assert "Synthèse des décaissements fournisseurs" in html
    assert ("Rapport " + "manager") not in html
    assert "Prévision mensuelle" in html
    assert "Réinitialiser" in html
    assert "Échéance" in html
    assert "Généré le" in html
    assert "kpiTotal" in html
    assert "Nombre de factures" in html
    assert "monthFilter" in html
    assert "supplierFilter" in html
    assert "projectFilter" in html
    assert "resetBtn" in html
    assert "runtimeError" in html
    assert "const el = {" in html
    assert js.count("data=JSON.parse") == 1
    assert js.count("window.addEventListener('resize',render)") == 1
    assert html.count("charts={}") == 0
    assert js.count("function fillSelect(") == 1
    assert js.count("function activeRows(") == 1
    assert js.count("function compatibleRows(") == 1
    assert js.count("function searchMatches(") == 1
    assert js.count("function compatibleValues(") == 1
    assert js.count("function clearProjectIfIncompatible(") == 1
    assert js.count("function clearSupplierIfIncompatible(") == 1
    assert js.count("function clearMonthDependentFilters(") == 1
    assert js.count("function clearIncompatibleFilters(") == 1
    assert js.count("function updateCascadingFilters(") == 1
    assert js.count("function resetFilters(") == 1
    assert js.count("function parsePayload(") == 1
    assert js.count("function money(") == 1
    assert js.count("function unique(") == 1
    assert js.count("function sumBy(") == 1
    assert js.count("function topEntries(") == 1
    assert js.count("function recordText(") == 1
    assert js.count("function dataset(") == 1
    assert js.count("function roundRect(") == 1
    assert js.count("function drawChart(") == 1
    assert js.count("function computeInsights(") == 1
    assert js.count("function renderInsights(") == 1
    assert js.count("function renderKpis(") == 1
    assert js.count("function renderSupplier(") == 1
    assert js.count("function renderFilterChips(") == 1
    assert js.count("function renderTable(") == 1
    assert js.count("function render(){") == 1
    assert js.count("function init(){") == 1
    assert js.count("function start(){") == 1
    assert js.count("try{init()}catch") == 1
    assert js.count("try{init()}catch(error)") == 1
    assert "document.addEventListener('DOMContentLoaded',start,{once:true})" in js
    assert "console.log('data loaded')" in js
    assert "console.log('row count',records.length)" in js
    assert "console.log('render start')" in js
    assert "console.log('render success')" in js
    assert "window.onerror" in js
    assert "window.onunhandledrejection" in js
    assert "getContext('2d')" in js
    assert "generated.textContent=" not in html
    assert "sourceFilters.innerHTML" not in html
    assert "monthFilter.value" not in html
    assert "supplierFilter.value" not in html
    assert "projectFilter.value" not in html
    assert "searchBox.value" not in html
    assert "tableSearch.value" not in html
    assert "resultCount.textContent" not in html
    assert "activeChips.innerHTML" not in html
    assert "resetBtn.onclick" not in html
    assert "csvBtn.onclick" not in html
    assert "Object.values" not in html
    assert "replaceAll" not in html
    assert "?." not in html
    assert "??" not in html
    assert "cdn.jsdelivr" not in html
    assert "new Chart" not in html
    assert "http://" not in html
    assert "https://" not in html
    assert "csvBtn" in html
    assert "pageSize" in html
    assert "tableSearch" in html
    assert "tableChips" in html
    assert "noResults" in html
    assert "noResultsReset" in html
    assert "Aucune facture ne correspond aux filtres sélectionnés." in html
    assert "Réinitialiser les filtres" in html
    assert "Certains filtres incompatibles ont été réinitialisés." in html
    assert "streamlit" not in html.lower()


def test_interactive_html_report_javascript_has_no_duplicate_function_names() -> None:
    cleaned = clean_invoices(make_invoice_frame())
    forecast = build_monthly_forecast(cleaned)
    metrics = compute_executive_metrics(cleaned, forecast)
    kpis = compute_kpis(cleaned, forecast)
    insights = generate_executive_insights(cleaned, forecast, metrics)
    filters = {"suppliers": [], "codes": [], "project_names": [], "months": [], "min_amount": 0.0}

    html = build_interactive_html_report(filters, cleaned, kpis, metrics, insights).decode("utf-8")
    js = extract_html_report_js(html)
    names = re.findall(r"function\s+([A-Za-z_$][\w$]*)\s*\(", js)

    assert len(names) == len(set(names))
    for name in names:
        assert js.count(f"function {name}(") == 1
    assert html.count('<script id="report-data" type="application/json">') == 1
    assert html.count("</script>\n<script>") == 1
    assert html.count("</main>") == 1
    assert html.count("</body></html>") == 1
    assert html.count("const el =") == 1
    assert js.count("try{init()}catch(error)") == 1
    assert js.index("try{init()}catch(error)") > js.index("function start(){")


def test_interactive_html_report_has_cascading_filter_logic() -> None:
    cleaned = clean_invoices(make_invoice_frame())
    forecast = build_monthly_forecast(cleaned)
    metrics = compute_executive_metrics(cleaned, forecast)
    kpis = compute_kpis(cleaned, forecast)
    insights = generate_executive_insights(cleaned, forecast, metrics)
    filters = {"suppliers": [], "codes": [], "project_names": [], "months": [], "min_amount": 0.0}

    html = build_interactive_html_report(filters, cleaned, kpis, metrics, insights).decode("utf-8")
    js = extract_html_report_js(html)

    assert "function compatibleRows(" in js
    assert "function searchMatches(" in js
    assert "function compatibleValues(" in js
    assert "function clearProjectIfIncompatible(" in js
    assert "function clearSupplierIfIncompatible(" in js
    assert "function clearMonthDependentFilters(" in js
    assert "function clearIncompatibleFilters(" in js
    assert "function updateCascadingFilters(" in js
    assert "function resetFilters(" in js
    assert "fillSelect(el.projectFilter,compatibleValues('project_code','project'))" in js
    assert "fillSelect(el.supplierFilter,compatibleValues('supplier','supplier'))" in js
    assert "fillSelect(el.monthFilter,compatibleValues('month','month'))" in js
    assert "clearProjectIfIncompatible();updateCascadingFilters(true);render()" in js
    assert "clearSupplierIfIncompatible();updateCascadingFilters(true);render()" in js
    assert "clearMonthDependentFilters();updateCascadingFilters(true);render()" in js
    assert "el.projectFilter['value']=''" in js
    assert "el.supplierFilter['value']=''" in js
    assert "Certains filtres incompatibles ont été réinitialisés." in js
    assert "Aucune facture ne correspond aux filtres sélectionnés." in html
    assert "Réinitialiser les filtres" in html
    assert "tableSearch" in html
    assert "http://" not in html
    assert "https://" not in html


def test_interactive_html_cascading_logic_uses_compatible_rows_for_small_supplier_project_pairs() -> None:
    raw = pd.DataFrame(
        {
            "Nom du fournisseur": ["Supplier A", "Supplier B"],
            "Numéro réf. fournisseur": ["A-1", "B-1"],
            "Date d'échéance": ["15/01/2025", "20/01/2025"],
            "Montant": [1000, 2000],
            "Date comptable": ["01/01/2025", "02/01/2025"],
            "Code Affaire": ["X", "Y"],
            "Nom Projet": ["Project X", "Project Y"],
            "MOIS": ["01", "01"],
        }
    )
    cleaned = clean_invoices(raw)
    forecast = build_monthly_forecast(cleaned)
    metrics = compute_executive_metrics(cleaned, forecast)
    kpis = compute_kpis(cleaned, forecast)
    insights = generate_executive_insights(cleaned, forecast, metrics)
    filters = {"suppliers": [], "codes": [], "project_names": [], "months": [], "min_amount": 0.0}

    html = build_interactive_html_report(filters, cleaned, kpis, metrics, insights).decode("utf-8")
    js = extract_html_report_js(html)

    assert "Supplier A" in html
    assert "Project X / X" in html
    assert "Supplier B" in html
    assert "Project Y / Y" in html
    assert "function compatibleRows(" in js
    assert "skipKey==='supplier'" in js
    assert "skipKey==='project'" in js
    assert "fillSelect(el.projectFilter,compatibleValues('project_code','project'))" in js
    assert "fillSelect(el.supplierFilter,compatibleValues('supplier','supplier'))" in js
    assert "searchMatches(record)" in js
    assert "(el.searchBox['value']+' '+el.tableSearch['value'])" not in js
    assert "const search=" not in js


def test_interactive_html_runtime_error_message_uses_escaped_newline() -> None:
    cleaned = clean_invoices(make_invoice_frame())
    forecast = build_monthly_forecast(cleaned)
    metrics = compute_executive_metrics(cleaned, forecast)
    kpis = compute_kpis(cleaned, forecast)
    insights = generate_executive_insights(cleaned, forecast, metrics)
    filters = {"suppliers": [], "codes": [], "project_names": [], "months": [], "min_amount": 0.0}

    html = build_interactive_html_report(filters, cleaned, kpis, metrics, insights).decode("utf-8")
    js = extract_html_report_js(html)

    assert js.count("function showRuntimeError(") == 1
    assert js.count("function exportCsv(") == 1
    assert "rapport interactif.\n'+message" not in js
    assert "lines.join('\n')" not in js
    assert "rapport interactif.\\n" in js or ".join('\\n')" in js


def test_interactive_html_embeds_valid_json_with_special_text_values() -> None:
    raw = pd.DataFrame(
        {
            "Nom du fournisseur": ['Société "L\'Atlas"\nNord'],
            "Numéro réf. fournisseur": ['REF-"A\'1"'],
            "Date d'échéance": [46139],
            "Montant": [1250.75],
            "Date comptable": ["27.01.2026"],
            "Code Affaire": ["CA-ÉNERGIE"],
            "Nom Projet": ['Projet "O\'Neil"\nÉlectricité <Phase 1>'],
            "MOIS": ["04"],
        }
    )
    cleaned = clean_invoices(raw)
    forecast = build_monthly_forecast(cleaned)
    metrics = compute_executive_metrics(cleaned, forecast)
    kpis = compute_kpis(cleaned, forecast)
    insights = generate_executive_insights(cleaned, forecast, metrics)
    filters = {"suppliers": [], "codes": [], "project_names": [], "months": [], "min_amount": 0.0}

    html = build_interactive_html_report(filters, cleaned, kpis, metrics, insights).decode("utf-8")
    payload_text = html.split('<script id="report-data" type="application/json">', 1)[1].split("</script>", 1)[0]
    payload = json.loads(payload_text)

    assert payload["records"][0]["supplier"] == 'Société "L\'Atlas"\nNord'
    assert payload["records"][0]["project"] == 'Projet "O\'Neil"\nÉlectricité <Phase 1>'
    assert payload["records"][0]["project_code"] == 'Projet "O\'Neil"\nÉlectricité <Phase 1> / CA-ÉNERGIE'
    assert "cdn.jsdelivr" not in html
    assert "http://" not in html
    assert "https://" not in html


def test_manager_chart_dataset_generation() -> None:
    records = [
        {"month": "2026-04", "supplier": "A", "project_code": "P1", "amount": 100.0},
        {"month": "2026-04", "supplier": "B", "project_code": "P2", "amount": 50.0},
        {"month": "2026-05", "supplier": "A", "project_code": "P1", "amount": 25.0},
    ]

    datasets = manager_chart_datasets(records)

    assert datasets["monthly"][0] == {"label": "2026-04", "amount": 150.0}
    assert datasets["suppliers"][0] == {"label": "A", "amount": 125.0}
    assert datasets["projects"][0] == {"label": "P1", "amount": 125.0}


def test_manager_supplier_detail_calculations() -> None:
    records = [
        {"supplier": "A", "project_code": "P1", "amount": 100.0, "due_date": "2026-04-10", "accounting_date": "2026-03-01"},
        {"supplier": "A", "project_code": "P2", "amount": 50.0, "due_date": "2026-04-20", "accounting_date": "2026-03-05"},
        {"supplier": "B", "project_code": "P3", "amount": 25.0, "due_date": "2026-05-01", "accounting_date": "2026-04-01"},
    ]

    overall = manager_supplier_detail(records)
    supplier = manager_supplier_detail(records, "A")

    assert overall["supplier_count"] == 2
    assert overall["project_count"] == 3
    assert overall["highest_invoice"] == 100.0
    assert supplier["invoice_count"] == 2
    assert supplier["total_amount"] == 150.0
    assert supplier["latest_due_date"] == "2026-04-20"


def test_manager_dynamic_insight_generation() -> None:
    records = [
        {"month": "2026-04", "supplier": "A", "project_code": "P1", "amount": 900.0},
        {"month": "2026-05", "supplier": "B", "project_code": "P2", "amount": 100.0},
    ]

    insights = manager_dynamic_insights(records)

    assert any(insight["level"] in {"WARNING", "CRITICAL"} for insight in insights)
    assert any("2026-04" in insight["text"] for insight in insights)


def test_manager_csv_generation() -> None:
    records = [
        {"supplier": "A", "reference": "R1", "due_date": "2026-04-27", "amount": 100.0, "accounting_date": "2026-01-27", "code": "C1", "project": "P1", "month": "2026-04"}
    ]

    csv = manager_records_to_csv(records)

    assert csv.startswith("supplier;reference;due_date;amount")
    assert "2026-04-27" in csv
    assert "100.0" in csv
